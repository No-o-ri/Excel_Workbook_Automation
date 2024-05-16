
import openpyxl as xl
from openpyxl.chart import LineChart, Reference


def process_workbook_data(file_name, interest, compound, time_duration):
    wb = xl.load_workbook(file_name)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        interest_price = 0.0
        if cell.value is None:
            continue  # Skip empty cells
        else:
            interest_price = cell.value * ((1 + interest / compound) ** (compound * time_duration))
        # interest_price = cell.value * pow((1-interest/compound), compound * time_duration)
        interest_cell = sheet.cell(row, 4)
        interest_cell.value = interest_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = LineChart()
    chart.title = 'Interest Rates At Different Principal Prices'
    chart.style = 8
    chart.y_axis.title = "Price"
    chart.x_axis.title = "Time"
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(file_name)
